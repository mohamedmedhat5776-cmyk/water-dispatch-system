from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import pandas as pd
import openpyxl
from datetime import datetime
import os

app = Flask(__name__)
CORS(app)

class ExcelHandler:
    def __init__(self, file_path):
        self.file_path = file_path
        # الأسماء الصحيحة كما في ملف Excel
        self.dispatch_sheet = " Daily Dispatch"  # مسافة في الأول
        self.water_quantity_sheet = "Water Quantity" 
        self.monthly_production_sheet = " Monthly production"  # مسافة في الأول
        self.second_meter_sheet = "Second meter production"  # مسافة في الأول
        
        print(f"📊 Excel file path: {file_path}")
        print(f"📊 File exists: {os.path.exists(file_path)}")
        
        # تحقق من وجود الـ sheets
        try:
            workbook = openpyxl.load_workbook(self.file_path)
            print(f"📊 Available sheets: {workbook.sheetnames}")
            workbook.close()
        except Exception as e:
            print(f"❌ Error checking sheets: {e}")
    
    def update_dispatch_data(self, location, quantity, day_of_month):
        """تحديث بيانات التوزيع اليومي"""
        try:
            print(f"📍 Updating dispatch: '{location}', Qty: {quantity}, Day: {day_of_month}")
            
            workbook = openpyxl.load_workbook(self.file_path)
            
            if self.dispatch_sheet not in workbook.sheetnames:
                print(f"❌ Sheet '{self.dispatch_sheet}' not found!")
                return False
            
            sheet = workbook[self.dispatch_sheet]
            
            # البحث عن الصف المناسب للموقع
            row_num = None
            for row in range(4, 80):
                cell_value = sheet.cell(row=row, column=2).value
                if cell_value and str(cell_value).strip() == location:
                    row_num = row
                    break
            
            print(f"🔍 Found row for '{location}': {row_num}")
            
            if row_num:
                # تحديد العمود بناءً على اليوم
                column_num = 6 + int(day_of_month)  # G=7 هو اليوم 1
                print(f"📝 Updating column: {column_num} (Day {day_of_month})")
                
                # تحديث الخلية
                sheet.cell(row=row_num, column=column_num).value = float(quantity)
                
                # تحديث العمود AL (Total quantity dispatched)
                total_formula = f"=SUM(G{row_num}:AK{row_num})"
                sheet.cell(row=row_num, column=38).value = total_formula
                
                # تحديث العمود AM (Balance quantity)
                balance_formula = f"=E{row_num}-AL{row_num}"
                sheet.cell(row=row_num, column=39).value = balance_formula
                
                workbook.save(self.file_path)
                print("✅ Dispatch data saved successfully!")
                return True
            else:
                print(f"❌ Location '{location}' not found in Excel")
                return False
            
        except Exception as e:
            print(f"❌ Error updating dispatch data: {e}")
            return False
    
    def update_water_quantity(self, ship_number, meter1_final, meter1_previous, date):
        """تحديث بيانات كمية المياه (عداد 1)"""
        try:
            print(f"🚢 Updating water quantity - Ship: {ship_number}")
            print(f"🔢 Meter1 Final: {meter1_final}, Meter1 Previous: {meter1_previous}")
            
            workbook = openpyxl.load_workbook(self.file_path)
            
            if self.water_quantity_sheet not in workbook.sheetnames:
                print(f"❌ Sheet '{self.water_quantity_sheet}' not found!")
                return False
            
            sheet = workbook[self.water_quantity_sheet]
            
            # تحديد الصف بناءً على رقم السفينة
            row_num = 6 + int(ship_number)  # الصفوف من 7 إلى 10
            print(f"📊 Water quantity row: {row_num}")
            
            # تحديث القراءة النهائية لليوم الحالي (العمود E)
            sheet.cell(row=row_num, column=5).value = float(meter1_final)
            print(f"✅ Updated Final Reading (Column E): {meter1_final}")
            
            # تحديث القراءة الأولية من اليوم السابق (العمود D)
            sheet.cell(row=row_num, column=4).value = float(meter1_previous)
            print(f"✅ Updated Initial Reading (Column D): {meter1_previous}")
            
            # حساب وحفظ الحجم تلقائياً
            volume = float(meter1_final) - float(meter1_previous)
            sheet.cell(row=row_num, column=6).value = volume
            print(f"📈 Volume calculated: {volume}")
            
            workbook.save(self.file_path)
            print("✅ Water quantity saved successfully!")
            return True
            
        except Exception as e:
            print(f"❌ Error updating water quantity: {e}")
            return False
    
    def update_monthly_production(self, ship_number, meter1_final, date):
        """تحديث الإنتاج الشهري (عداد 1)"""
        try:
            print(f"📅 Updating monthly production - Ship: {ship_number}, Meter1: {meter1_final}")
            
            workbook = openpyxl.load_workbook(self.file_path)
            
            if self.monthly_production_sheet not in workbook.sheetnames:
                print(f"❌ Sheet '{self.monthly_production_sheet}' not found!")
                return False
            
            sheet = workbook[self.monthly_production_sheet]
            
            # الحصول على اليوم من التاريخ
            day = datetime.strptime(date, '%Y-%m-%d').day
            print(f"📅 Day of month: {day}")
            
            # تحديد الصف بناءً على اليوم
            row_num = 8 + day  # الصفوف من 9
            print(f"📊 Monthly production row: {row_num}")
            
            # تحديد الأعمدة بناءً على رقم السفينة
            columns = {'1': 3, '2': 5, '3': 7, '4': 9}
            column_num = columns[ship_number]
            
            # تحديث القراءة النهائية لليوم الحالي
            sheet.cell(row=row_num, column=column_num).value = float(meter1_final)
            print(f"✅ Updated Monthly Production - Ship {ship_number}: {meter1_final}")
            
            workbook.save(self.file_path)
            print("✅ Monthly production saved successfully!")
            return True
            
        except Exception as e:
            print(f"❌ Error updating monthly production: {e}")
            return False
    
    def update_second_meter(self, ship_number, meter2_final, date):
        """تحديث عداد الإنتاج الثاني"""
        try:
            print(f"🔢 Updating second meter - Ship: {ship_number}, Meter2: {meter2_final}")
            
            workbook = openpyxl.load_workbook(self.file_path)
            
            if self.second_meter_sheet not in workbook.sheetnames:
                print(f"❌ Sheet '{self.second_meter_sheet}' not found!")
                return False
            
            sheet = workbook[self.second_meter_sheet]
            
            # الحصول على اليوم من التاريخ
            day = datetime.strptime(date, '%Y-%m-%d').day
            print(f"📅 Day of month: {day}")
            
            # تحديد الصف بناءً على اليوم
            row_num = 8 + day  # الصفوف من 9
            print(f"📊 Second meter row: {row_num}")
            
            # تحديد الأعمدة بناءً على رقم السفينة
            columns = {'1': 3, '2': 5, '3': 7, '4': 9}
            column_num = columns[ship_number]
            
            # تحديث القراءة النهائية لعداد 2
            sheet.cell(row=row_num, column=column_num).value = float(meter2_final)
            print(f"✅ Updated Second Meter - Ship {ship_number}: {meter2_final}")
            
            workbook.save(self.file_path)
            print("✅Second meter saved successfully!")
            return True
            
        except Exception as e:
            print(f"❌ Error updating second meter: {e}")
            return False

# المسار الكامل لملف Excel
excel_file_path = r"E:\musandam\New Water bot\Dispatch order.xlsx"
excel_handler = ExcelHandler(excel_file_path)

@app.route('/')
def index():
    return send_from_directory('.', 'index.html')

@app.route('/save_data', methods=['POST'])
def save_data():
    try:
        data = request.json
        print(f"📨 Received data: {data}")
        
        if data['type'] == 'dispatch':
            success = excel_handler.update_dispatch_data(
                data['location'],
                data['quantity'],
                data['dayOfMonth']
            )
        elif data['type'] == 'meter':
            # تحديث بيانات العدادات
            success1 = excel_handler.update_water_quantity(
                data['shipNumber'],
                data['meter1Final'],      # عداد 1 لليوم الحالي
                data['meter1Previous'],   # عداد 1 لليوم السابق
                data['date']
            )
            success2 = excel_handler.update_monthly_production(
                data['shipNumber'],
                data['meter1Final'],      # عداد 1 لليوم الحالي
                data['date']
            )
            success3 = excel_handler.update_second_meter(
                data['shipNumber'],
                data['meter2Final'],      # عداد 2 لليوم الحالي
                data['date']
            )
            success = success1 and success2 and success3
        else:
            success = False
            
        return jsonify({'success': success, 'message': 'تم الحفظ بنجاح' if success else 'فشل في الحفظ'})
        
    except Exception as e:
        print(f"🔥 Error in save_data: {e}")
        return jsonify({'success': False, 'message': f'خطأ: {str(e)}'})

if __name__ == '__main__':
    print("🚀 Starting Water Dispatch Application...")
    print("🌐 Local: http://localhost:5000")
    
    import socket
    hostname = socket.gethostname()
    local_ip = socket.gethostbyname(hostname)
    
    print(f"🌍 Network: http://{local_ip}:5000")
    print("📱 Available on all devices in your network!")
    print("🔧 If not working, check Windows Firewall settings!")
    
    # إضافة threaded لتحسين الأداء
    app.run(debug=True, host='0.0.0.0', port=5000, threaded=True)